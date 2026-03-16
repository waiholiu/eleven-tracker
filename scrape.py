"""
Scrapes https://11clubhouse.com/eleven-online-players.html for online player counts
and appends a timestamped row to an Excel file on OneDrive via Microsoft Graph API.
"""

import os
import re
import json
import sys
from datetime import datetime, timezone

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from io import BytesIO

# ── Configuration from environment ──────────────────────────────────────────
TENANT_ID = os.environ["AZURE_TENANT_ID"]
CLIENT_ID = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
# Path inside the user's OneDrive, e.g. "ElevenTracker/players.xlsx"
ONEDRIVE_FILE_PATH = os.environ.get("ONEDRIVE_FILE_PATH", "ElevenTracker/players.xlsx")
# The user's principal name (email), e.g. "user@contoso.com"
ONEDRIVE_USER = os.environ["ONEDRIVE_USER"]

SCRAPE_URL = "https://11clubhouse.com/eleven-online-players.html"
GRAPH_BASE = "https://graph.microsoft.com/v1.0"


def get_access_token() -> str:
    """Obtain an access token using client credentials flow."""
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    return resp.json()["access_token"]


def scrape_players() -> dict:
    """Scrape the page and return total + per-country player counts."""
    resp = requests.get(SCRAPE_URL, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Total players from the <h2> heading, e.g. "262 players online right now !"
    h2 = soup.find("h2", class_="title_shadow")
    total_match = re.search(r"(\d+)\s+players?\s+online", h2.get_text() if h2 else "")
    total = int(total_match.group(1)) if total_match else 0

    # Country data from title attributes like 'China: 70 users online'
    countries = {}
    for link in soup.select("ol.list_pays li a.tooltipgo"):
        title = link.get("title", "")
        m = re.match(r"(.+?):\s*(\d+)\s+users?\s+online", title)
        if m:
            countries[m.group(1)] = int(m.group(2))

    return {"total": total, "countries": countries}


def download_workbook(token: str) -> Workbook | None:
    """Download existing Excel file from OneDrive, or return None if not found."""
    url = f"{GRAPH_BASE}/users/{ONEDRIVE_USER}/drive/root:/{ONEDRIVE_FILE_PATH}:/content"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    return load_workbook(BytesIO(resp.content))


def upload_workbook(token: str, wb: Workbook):
    """Upload the workbook back to OneDrive (creates or overwrites)."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    url = f"{GRAPH_BASE}/users/{ONEDRIVE_USER}/drive/root:/{ONEDRIVE_FILE_PATH}:/content"
    resp = requests.put(url, headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }, data=buf.read())
    resp.raise_for_status()


def build_row(timestamp: str, data: dict, all_countries: list[str]) -> list:
    """Build a spreadsheet row: [timestamp, total, country1, country2, ...]."""
    row = [timestamp, data["total"]]
    for c in all_countries:
        row.append(data["countries"].get(c, 0))
    return row


def main():
    print("Scraping player data...")
    data = scrape_players()
    print(f"  Total: {data['total']} players, {len(data['countries'])} countries")

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    print("Authenticating with Microsoft Graph...")
    token = get_access_token()

    print("Downloading existing spreadsheet...")
    wb = download_workbook(token)

    if wb is None:
        # Create new workbook with headers
        print("  No existing file found — creating new spreadsheet.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Player Counts"
        all_countries = sorted(data["countries"].keys())
        headers = ["Timestamp", "Total Players"] + all_countries
        ws.append(headers)
    else:
        ws = wb.active

    # Determine the country columns from the header row
    header_row = [cell.value for cell in ws[1]]
    existing_countries = header_row[2:]  # everything after "Timestamp" and "Total Players"

    # Add any new countries not yet in the header
    new_countries = sorted(set(data["countries"].keys()) - set(existing_countries))
    if new_countries:
        print(f"  Adding {len(new_countries)} new country columns: {new_countries}")
        for c in new_countries:
            existing_countries.append(c)
            ws.cell(row=1, column=len(existing_countries) + 2, value=c)

    all_countries = existing_countries
    row = build_row(timestamp, data, all_countries)
    ws.append(row)

    print("Uploading spreadsheet to OneDrive...")
    upload_workbook(token, wb)
    print(f"Done! Row added at {timestamp}")


if __name__ == "__main__":
    main()
