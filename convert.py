"""
Converts data/players.xlsx to data/players.json for the web frontend.
"""

import json
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH = Path("data/players.xlsx")
JSON_PATH = Path("data/players.json")


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    countries = headers[2:]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {
            "timestamp": row[0],
            "total": row[1],
            "countries": {},
        }
        for i, country in enumerate(countries):
            val = row[i + 2]
            if val and val > 0:
                entry["countries"][country] = val
        rows.append(entry)

    output = {
        "countries": sorted(countries),
        "data": rows,
    }

    JSON_PATH.write_text(json.dumps(output, indent=2))
    print(f"Wrote {len(rows)} rows to {JSON_PATH}")


if __name__ == "__main__":
    main()
